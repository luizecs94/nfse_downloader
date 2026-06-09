"""
Gerador de relatório Excel com notas fiscais encontradas no portal.

Usado quando o download automático não é possível (CAPTCHA no portal web).
Gera um arquivo .xlsx formatado com todas as notas, dados e links diretos.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

logger = logging.getLogger(__name__)

# Colunas e larguras do relatório
_COLUNAS = [
    ("Empresa",          28),
    ("CNPJ",             18),
    ("Número NFS-e",     14),
    ("Data Emissão",     14),
    ("Valor",            14),
    ("Status",           14),
    ("Chave de Acesso",  52),
    ("Link XML",         55),
    ("Link PDF/DANFSe",  55),
]

# Cores
_COR_CABECALHO_FUNDO   = "1F4E79"   # azul escuro
_COR_CABECALHO_TEXTO   = "FFFFFF"   # branco
_COR_LINHA_PAR         = "DCE6F1"   # azul claro
_COR_LINHA_IMPAR       = "FFFFFF"   # branco
_COR_AVISO_FUNDO       = "FFF2CC"   # amarelo claro
_COR_AVISO_TEXTO       = "7F6000"   # marrom escuro
_COR_BORDA             = "B8CCE4"


def gerar_relatorio(
    registros: List[Dict[str, Any]],
    caminho_saida: Path,
) -> Path:
    """
    Gera relatório Excel a partir de uma lista de registros de notas.

    Args:
        registros: Lista de dicts com chaves:
            empresa, cnpj, numero, data_emissao, valor, status,
            chave_acesso, link_xml, link_pdf
        caminho_saida: Caminho completo do arquivo .xlsx a gerar.

    Returns:
        Caminho do arquivo gerado.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "openpyxl não instalado. Execute:\n"
            "  .venv\\Scripts\\pip.exe install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "NFS-e Encontradas"

    # ------------------------------------------------------------------
    # Estilos reutilizáveis
    # ------------------------------------------------------------------
    fill_header = PatternFill("solid", fgColor=_COR_CABECALHO_FUNDO)
    fill_par    = PatternFill("solid", fgColor=_COR_LINHA_PAR)
    fill_impar  = PatternFill("solid", fgColor=_COR_LINHA_IMPAR)
    fill_aviso  = PatternFill("solid", fgColor=_COR_AVISO_FUNDO)

    font_header = Font(bold=True, color=_COR_CABECALHO_TEXTO, size=10)
    font_normal = Font(size=10)
    font_aviso  = Font(bold=True, color=_COR_AVISO_TEXTO, size=10)
    font_link   = Font(color="0563C1", underline="single", size=10)

    borda_fina = Side(style="thin", color=_COR_BORDA)
    borda = Border(
        left=borda_fina, right=borda_fina,
        top=borda_fina,  bottom=borda_fina,
    )
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esq    = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ------------------------------------------------------------------
    # Linha de aviso (linha 1)
    # ------------------------------------------------------------------
    ws.merge_cells("A1:I1")
    cel_aviso = ws["A1"]
    cel_aviso.value = (
        "⚠️  Download automático bloqueado por CAPTCHA do portal. "
        "Use os links abaixo para baixar manualmente cada nota."
    )
    cel_aviso.font      = font_aviso
    cel_aviso.fill      = fill_aviso
    cel_aviso.alignment = Alignment(horizontal="center", vertical="center")
    cel_aviso.border    = borda
    ws.row_dimensions[1].height = 20

    # ------------------------------------------------------------------
    # Cabeçalho (linha 2)
    # ------------------------------------------------------------------
    for col_idx, (titulo, largura) in enumerate(_COLUNAS, start=1):
        cel = ws.cell(row=2, column=col_idx, value=titulo)
        cel.font      = font_header
        cel.fill      = fill_header
        cel.alignment = alinhamento_centro
        cel.border    = borda
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    # ------------------------------------------------------------------
    # Dados
    # ------------------------------------------------------------------
    for i, reg in enumerate(registros):
        linha = i + 3
        fill  = fill_par if i % 2 == 0 else fill_impar

        valores = [
            reg.get("empresa",       ""),
            _formatar_cnpj(reg.get("cnpj", "")),
            reg.get("numero",        ""),
            reg.get("data_emissao",  ""),
            reg.get("valor",         ""),
            reg.get("status",        ""),
            reg.get("chave_acesso",  ""),
            reg.get("link_xml",      ""),
            reg.get("link_pdf",      ""),
        ]

        for col_idx, valor in enumerate(valores, start=1):
            cel = ws.cell(row=linha, column=col_idx, value=valor)
            cel.fill      = fill
            cel.border    = borda
            cel.alignment = alinhamento_esq

            # Colunas de link — aplica hiperlink se tiver URL
            if col_idx in (8, 9) and valor and valor.startswith("http"):
                cel.hyperlink = valor
                cel.font      = font_link
            else:
                cel.font = font_normal

        ws.row_dimensions[linha].height = 16

    # ------------------------------------------------------------------
    # Rodapé com totais
    # ------------------------------------------------------------------
    linha_rodape = len(registros) + 3
    ws.merge_cells(f"A{linha_rodape}:I{linha_rodape}")
    cel_total = ws[f"A{linha_rodape}"]
    cel_total.value = (
        f"Total: {len(registros)} nota(s) encontrada(s)  |  "
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    )
    cel_total.font      = Font(italic=True, size=9, color="595959")
    cel_total.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[linha_rodape].height = 14

    # ------------------------------------------------------------------
    # Salvar
    # ------------------------------------------------------------------
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    logger.info("Relatório salvo: %s", caminho_saida)
    return caminho_saida


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _formatar_cnpj(cnpj: str) -> str:
    """Formata CNPJ: 00.000.000/0000-00"""
    cnpj = "".join(c for c in str(cnpj) if c.isdigit())
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj
